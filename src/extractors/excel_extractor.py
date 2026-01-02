"""
Excel spreadsheet extractor using openpyxl and pandas.

Extracts data from .xlsx and .xls files, converting tabular
data into a structured text format suitable for LLM processing.
"""

from pathlib import Path
from typing import Any, ClassVar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.extractors.base import DocumentExtractor, ExtractionError
from src.models import ExtractedDocument


class ExcelExtractor(DocumentExtractor):
    """
    Extracts text content from Excel spreadsheets.

    Handles both .xlsx (modern) and .xls (legacy via pandas) formats.
    Converts tabular data to structured text with clear formatting.
    """

    SUPPORTED_EXTENSIONS: ClassVar[tuple[str, ...]] = (".xlsx", ".xls")

    def extract(self, file_path: Path) -> ExtractedDocument:
        """
        Extract text from an Excel file.

        Args:
            file_path: Path to the Excel file.

        Returns:
            ExtractedDocument with structured text representation.

        Raises:
            ExtractionError: If the file cannot be read.
        """
        self._validate_file(file_path)

        try:
            extension = file_path.suffix.lower()

            if extension == ".xlsx":
                content = self._extract_xlsx(file_path)
            else:  # .xls
                content = self._extract_xls(file_path)

            if not content.strip():
                raise ExtractionError("Spreadsheet contains no data", file_path)

            return self._create_result(content, file_path)

        except InvalidFileException as e:
            raise ExtractionError(
                "File is not a valid Excel document or is corrupted", file_path, cause=e
            ) from e
        except ExtractionError:
            raise
        except Exception as e:
            raise ExtractionError(f"Unexpected error: {e}", file_path, cause=e) from e

    def _extract_xlsx(self, file_path: Path) -> str:
        """
        Extract text from .xlsx using openpyxl.

        Args:
            file_path: Path to the .xlsx file.

        Returns:
            Text representation of all sheets.
        """
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        text_parts: list[str] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = self._extract_sheet(sheet, sheet_name)
            if sheet_text:
                text_parts.append(sheet_text)

        workbook.close()
        return "\n\n".join(text_parts)

    def _extract_xls(self, file_path: Path) -> str:
        """
        Extract text from .xls using pandas.

        Args:
            file_path: Path to the .xls file.

        Returns:
            Text representation of all sheets.
        """
        # pandas can read .xls files with xlrd engine
        excel_data: dict[str, pd.DataFrame] = pd.read_excel(
            file_path, sheet_name=None, engine="xlrd"
        )

        text_parts: list[str] = []

        for sheet_name, df in excel_data.items():
            if not df.empty:
                sheet_text = f"=== Sheet: {sheet_name} ===\n"
                sheet_text += df.to_string(index=False)
                text_parts.append(sheet_text)

        return "\n\n".join(text_parts)

    def _extract_sheet(self, sheet: Any, sheet_name: str) -> str:
        """
        Extract text from a single worksheet.

        Args:
            sheet: openpyxl worksheet object.
            sheet_name: Name of the sheet.

        Returns:
            Text representation of the sheet.
        """
        rows: list[str] = []

        for row in sheet.iter_rows(values_only=True):
            # Convert cell values to strings, handling None
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cell.strip() for cell in cells):  # Skip empty rows
                rows.append(" | ".join(cells))

        if rows:
            return f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows)
        return ""
